#IMPORTAMOS LAS LIBRERIAS NECESARIAS
import openpyxl
import os


#CREAMOS LA CLASE
class excell:
    #METODO CONSTRUCTOR
    def __init__(self):
        self.ruta = "static/contactos.xlsx"
    
    #METODO PARA CREAR EL ARCHIVO
    def crear_archivo(self):
        #CREAMOS UN NUEVO LIBRO
        libro = openpyxl.Workbook()

        #GUARDEMOS EL LIBRO EN LA RUTTA ESPEFICICADA
        libro.save(self.ruta)

    #METODO PARA VER EL ARCHIVO
    def abrir_archivo(self):
        #ABRIMOS EL ARCHIVIO DE EXCELL
        archivo = openpyxl.load_workbook(self.ruta)

        #SELECCIONAMOS LA HOJA DEL ARCHIVO
        hoja = archivo['Sheet']

        #HACEMOS ALGO CON LOS DATOS
        for fila in hoja.iter_rows(values_only=True):
            print(fila)

    #METODO PARA LLENAR LOS DATOS DEL ARCHIVO
    def llenar_archivo(self):
        #CREAMOS LA VARIABLE DE ARCHIVO
        archivo = openpyxl.load_workbook(self.ruta)
        #CREAMOS UNA VARIABLE CON LA HOJA QUE VAMOS A TTRABAJAR
        hoja = archivo["Sheet"]

        #AGREGAMOS LOS DATOS A LA HOJA
        datos = [
            ["Nombre", "Telefono"]
        ]

        #CREAMOS UN CICLO FOR QUE VA A PREGUNTAR CUANTOS CONTACTOS VA A AGREGAR
        contador = int(input("¿Cuantos contactos vas a guardar?: "))

        for i in range(contador):
            #DEFINIMOS LOS VALORES QUE VAMOS A GUARDAR EN EL ARREGLO DATOS
            nombre = input("Escribe el nombre del contacto: ")
            telefono = int(input("Escribe el numero del contacto: "))
            dato = [nombre , telefono]

            #UNA VEZ LLENOS LOS DATOS GUARDAMOS LOS DATOS EN EL ARREGLO PRINCIPAL
            datos.append(dato)
        
        #UNA VEZ LLENOS LOS DATOS AGREGAMOS TODO A LA HOJA
        for fila in datos:
            hoja.append(fila)
        
        #GUARDEMOS LOS CAMBIOS
        archivo.save(self.ruta)
    
    #METODO PARA ELIMINAR UNA HOJA DE CALCULO
    def eliminar_archivo(self):
        #CERRAMOS LOS ARHCIVOS DE EXCELL QUE ESTEN ABIERTOS
        os.system("taskkill /f /im excel.exe")

        #ELIMINA EL ARCHIVO
        os.remove(self.ruta)

#e = excell()
#e.eliminar_archivo()